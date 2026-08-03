import csv
import re
import unicodedata
from zipfile import BadZipFile
from pathlib import Path

from django.db import transaction

from .models import SUBJECT_GROUPS, Answer, Category, Document, Question, Subject


class ImportError(Exception):
    pass


EXCEL_DB_MAP = [
    ("Loại kiến thức (Chuyên môn/nghiệp vụ)", "category"),
    ("TT", "question_order"),
    ("QUESTION", "question.content"),
    ("A / 1", "answer A"),
    ("B / 2", "answer B"),
    ("C / 3", "answer C"),
    ("D / 4", "answer D"),
    ("ANS", "correct answer (A-D or 1-4)"),
    (
        "rating / RATING",
        "subject group(s) APS/ADC/ACC HAN/SUP/SUP ACS HAN/ACS SUP HCM (optional; multi allowed e.g. SUP,APS)",
    ),
]

CATEGORY_ALIASES = {
    "AIS": "Aeronautical Information Services",
    "AIS:": "Aeronautical Information Services",
    "ATC FACILITIES": "ATC Facilities",
    "FACILITIES": "ATC Facilities",
    "GENERAL KNOWLEDGE": "General Knowledge",
    "GENENRAL KNOWLEDGE": "General Knowledge",
    "HUMAN": "Human Factors",
    "METEOLOGY": "Meteorology",
    "METEOROLOGY": "Meteorology",
    "METEO": "Meteorology",
    "NAV": "Navigation",
    "NAVIGATION-PRINCIPLE": "Navigation",
    "OPERATION": "Operational Procedures",
    "OPERATION SYSTEM": "Operational Procedures",
    "OPERATIONAL PROCEDURE": "Operational Procedures",
    "SERVEILANCE SYSTEM": "Surveillance",
    "SURVEILLANCE SYSTEM": "Surveillance",
    "CHARACTERISTIC": "Aircraft Characteristics",
    "CHARACTERISTICS": "Aircraft Characteristics",
    "CAVOKS MEANS:": "Meteorology",
    "SHIFTS MANAGEMENT": "Shifts Management",
    "AIR TRAFFIC FLOWS": "Air Traffic Flow Management",
}

CATEGORY_KEYWORDS = [
    ("Aeronautical Information Services", ("ais", "aip", "notam", "snowtam", "ashtam", "aeronautical information")),
    ("Meteorology", ("weather", "metar", "taf", "cavok", "visibility", "cloud", "ceiling", "rvr", "qnh", "qfe", "wind", "fog", "mist", "thunderstorm", "meteorological", "turbulence", "icing")),
    ("Emergency and SAR", ("emergency", "distress", "urgency", "mayday", "pan-pan", "sar", "search and rescue", "fire", "hijack", "unlawful")),
    ("Surveillance", ("radar", "surveillance", "ssr", "transponder", "squawk", "mode c", "msaw", "vector")),
    ("Navigation", ("navigation", "vor", "dme", "ndb", "ils", "localizer", "glide", "rnav", "rnp", "waypoint", "bearing", "radial", "magnetic", "track")),
    ("Aerodrome", ("aerodrome", "airport", "runway", "taxiway", "apron", "tora", "toda", "asda", "lda", "papi", "vasi", "resa", "threshold")),
    ("Airspace Structure", ("airspace", "tma", "ctr", "cta", "prohibited", "restricted", "danger area", "airway", "rvsm", "flight level")),
    ("Air Traffic Flow Management", ("flow", "atfm", "slot", "capacity")),
    ("Human Factors", ("human", "fatigue", "stress", "teamwork", "crm")),
    ("Shifts Management", ("shift", "roster", "handover")),
    ("ATC Facilities", ("facility", "facilities", "frequency", "channel", "communication", "vhf", "uhf")),
    ("Regulations", ("law", "decree", "circular", "annex", "icao", "doc ", "regulation", "minister")),
    ("Operational Procedures", ("clearance", "separation", "holding", "approach", "departure", "arrival", "taxi", "ifr", "vfr", "control", "flight plan", "transfer of control")),
]

CATEGORY_RULES = [
    (category, ", ".join(keywords)) for category, keywords in CATEGORY_KEYWORDS
]


def _normalize(v):
    if v is None:
        return ""
    return str(v).strip()


def _compact_spaces(v):
    return re.sub(r"\s+", " ", _normalize(v)).strip()


def _header_key(v):
    text = _normalize(v)
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"\s+", " ", text).strip().upper()
    return text


def classify_question_category(topic, question):
    topic = _compact_spaces(topic)
    topic_key = _header_key(topic)
    if topic_key and topic_key not in {"KHAC", "OTHER", "NONE", "N/A"}:
        return CATEGORY_ALIASES.get(topic_key, topic)

    haystack = _header_key(f"{topic} {question}").lower()
    for category, keywords in CATEGORY_KEYWORDS:
        if any(keyword in haystack for keyword in keywords):
            return category
    return "General Knowledge"


def _classify_category(topic, question):
    return classify_question_category(topic, question)


def _source_prefix(path: Path):
    text = unicodedata.normalize("NFD", path.stem)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^A-Za-z0-9]+", "-", text).strip("-").upper()
    return text[:40] or "IMPORT"


def _source_tokens(path: Path):
    return set(_source_prefix(path).split("-"))


def _canonical_rating_token(token):
    """Map a rating token/alias to a SUBJECT_GROUPS code."""
    token = re.sub(r"[\s_\-]+", " ", _normalize(token).upper()).strip()
    if not token:
        return ""
    if token in SUBJECT_GROUPS:
        return token

    aliases = {
        "ACC": "ACC HAN",
        "ACCHAN": "ACC HAN",
        "ACC HAN": "ACC HAN",
        "ACC HN": "ACC HAN",
        "ACCHN": "ACC HAN",
        "SUP ACS HAN": "SUP ACS HAN",
        "SUPACSHAN": "SUP ACS HAN",
        "SUP_ACS_HAN": "SUP ACS HAN",
        "SUP ACS": "SUP ACS HAN",
        "ACS SUP HCM": "ACS SUP HCM",
        "ACSSUPHCM": "ACS SUP HCM",
        "ACS_SUP_HCM": "ACS SUP HCM",
        "ACS SUP HCM·": "ACS SUP HCM",
    }
    if token in aliases and aliases[token] in SUBJECT_GROUPS:
        return aliases[token]

    compact = token.replace(" ", "")
    for code in SUBJECT_GROUPS:
        if code.replace(" ", "") == compact:
            return code
    if compact in aliases and aliases[compact] in SUBJECT_GROUPS:
        return aliases[compact]
    return ""


def _parse_rating_values(value):
    """Parse one or many supported ratings from a cell.

    SUP questions may also belong to APS and/or ADC, so values like
    ``SUP,APS``, ``ADC/SUP``, or ``APS ADC SUP`` are supported.
    ``ACC HAN`` / ``SUP ACS HAN`` / ``ACS SUP HCM`` are dedicated banks.
    """
    text = _normalize(value).upper()
    if not text:
        return []

    found = []
    # Prefer comma/slash separators so multi-word groups stay intact.
    chunks = re.split(r"[,;/|]+", text)
    for chunk in chunks:
        chunk = re.sub(r"[\s_\-]+", " ", chunk).strip()
        if not chunk:
            continue
        code = _canonical_rating_token(chunk)
        if code:
            if code not in found:
                found.append(code)
            continue
        # Longest multi-word SUBJECT_GROUPS first (e.g. SUP ACS HAN before SUP).
        remainder = f" {chunk} "
        for code in sorted(SUBJECT_GROUPS, key=lambda c: len(c), reverse=True):
            needle = f" {code} "
            if needle in remainder:
                if code not in found:
                    found.append(code)
                remainder = remainder.replace(needle, " ")
        for token in remainder.split():
            code = _canonical_rating_token(token)
            if code and code not in found:
                found.append(code)
    return found


def _normalize_rating(value):
    ratings = _parse_rating_values(value)
    return ratings[0] if ratings else ""


def _normalize_answer_labels(value):
    """Normalize ANS values such as A/B/C/D, AC, or 1/2/3/4."""
    ans_raw = _normalize(value).upper().replace(" ", "")
    if not ans_raw:
        return ""
    if re.fullmatch(r"[1-4]+", ans_raw):
        conv = {"1": "A", "2": "B", "3": "C", "4": "D"}
        return "".join(conv[ch] for ch in ans_raw)
    return "".join(ch for ch in ans_raw if ch in "ABCD")


def _resolve_subject_targets(path: Path, subject_codes, rating_explicit: bool):
    """Route each question to subject DB(s).

    Priority:
    1) Explicit rating cell(s) in file (all SUBJECT_GROUPS; multi allowed)
    2) Form/default group chosen at import time
    3) Legacy filename heuristics (only when form group is APS/ADC)
    """
    if isinstance(subject_codes, str):
        codes = [subject_codes] if subject_codes in SUBJECT_GROUPS else []
    else:
        codes = [c for c in (subject_codes or []) if c in SUBJECT_GROUPS]

    order = {code: i for i, code in enumerate(SUBJECT_GROUPS)}
    dedicated = {"SUP", "ACC HAN", "SUP ACS HAN", "ACS SUP HCM"}

    if rating_explicit and codes:
        return sorted(dict.fromkeys(codes), key=lambda c: order.get(c, 99))

    if not codes:
        codes = ["APS"]

    # Dedicated banks must not be overridden by LTC/GCU/ACC filename heuristics.
    if dedicated.intersection(codes):
        return sorted(dict.fromkeys(codes), key=lambda c: order.get(c, 99))

    tokens = _source_tokens(path)
    if "ACSSUPHCM" in "".join(tokens) or (
        "ACS" in tokens and "SUP" in tokens and "HCM" in tokens
    ):
        return ["ACS SUP HCM"]
    if "SUPACS" in "".join(tokens) or (
        "SUP" in tokens and "ACS" in tokens and "HAN" in tokens
    ):
        return ["SUP ACS HAN"]
    if ("ACC" in tokens and "HAN" in tokens) or "ACCHAN" in tokens or "ACC" in tokens:
        return ["ACC HAN"]
    if "LTC" in tokens and set(codes) <= {"APS", "ADC"}:
        return ["ADC", "APS"]
    if "GCU" in tokens and set(codes) <= {"APS", "ADC"}:
        return ["ADC"]
    if "SUP" in tokens and set(codes) <= {"APS", "ADC"}:
        # Filename mentions SUP but form picked APS/ADC and file has no rating:
        # keep form choice; do not silently move bank to SUP.
        return sorted(dict.fromkeys(codes), key=lambda c: order.get(c, 99))

    return sorted(dict.fromkeys(codes), key=lambda c: order.get(c, 99))


def _target_subject_codes(path: Path, fallback_subject_code: str):
    # Backward-compatible helper for callers/tests without per-row rating flags.
    return _resolve_subject_targets(path, [fallback_subject_code], rating_explicit=False)


def _read_csv(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def _csv_field(row, *names):
    by_key = {_header_key(k): v for k, v in row.items() if k is not None}
    for name in names:
        key = _header_key(name)
        if key in by_key:
            return by_key[key]
    return ""


def _extract_from_table_rows(rows, default_subject_code):
    if not rows:
        return []

    # detect header row by QUESTION(S) + ANS
    header_idx = None
    for i, row in enumerate(rows):
        keys = {_header_key(x) for x in row if _normalize(x)}
        has_question = bool(keys & {"QUESTION", "QUESTIONS", "CAU HOI", "CÂU HỎI"})
        has_ans = bool(keys & {"ANS", "ANSWER", "DAP AN", "ĐÁP ÁN"})
        if has_question and has_ans:
            header_idx = i
            break
    if header_idx is None:
        raise ImportError("Khong tim thay dong header co QUESTION(S)/ANS.")

    headers = [_normalize(x) for x in rows[header_idx]]
    index = {_header_key(h): i for i, h in enumerate(headers) if h}

    def col(*names):
        for n in names:
            key = _header_key(n)
            if key in index:
                return index[key]
        return None

    c_topic = col(
        "LOẠI KIẾN THỨC (CHUYÊN MÔN/NGHIỆP VỤ)",
        "LOẠI KIẾN THỨC",
        "LOẠI KIẾN THỨC CHUYÊN MÔN",
        "LOAI KIEN THUC (CHUYEN MON/NGHIEP VU)",
        "LOAI KIEN THUC",
        "LOAI KIEN THUC CHUYEN MON",
    )
    c_tt = col("TT", "STT", "TT GOC", "TT GỐC")
    c_question = col("QUESTION", "QUESTIONS", "CAU HOI", "CÂU HỎI", "CAU HOI")
    # Answer option headers may be A/B/C/D or numeric 1/2/3/4.
    c_a = col("A", "1")
    c_b = col("B", "2")
    c_c = col("C", "3")
    c_d = col("D", "4")
    c_ans = col("ANS", "ANSWER", "DAP AN", "ĐÁP ÁN")
    # Optional column: rating / RATING (APS, ADC, ACC HAN, SUP).
    c_rating = col("RATING", "rating", "NHOM", "NHÓM", "SUBJECT")

    if c_question is None or c_ans is None or c_a is None or c_b is None:
        raise ImportError("Thieu cot bat buoc: QUESTION(S)/ANS va A/B (hoac 1/2).")

    extracted = []
    current_topic = ""
    for ridx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        question = _normalize(row[c_question] if c_question < len(row) else "")
        if not question:
            continue

        explicit_topic = _normalize(row[c_topic] if c_topic is not None and c_topic < len(row) else "")
        if explicit_topic:
            current_topic = explicit_topic
        topic = _classify_category(explicit_topic or current_topic or "Khac", question)
        tt = _normalize(row[c_tt] if c_tt is not None and c_tt < len(row) else "")
        raw_rating = _normalize(row[c_rating] if c_rating is not None and c_rating < len(row) else "")
        explicit_ratings = _parse_rating_values(raw_rating)
        rating_explicit = bool(explicit_ratings)
        subject_codes = explicit_ratings or [default_subject_code]

        ans_labels = _normalize_answer_labels(row[c_ans] if c_ans < len(row) else "")
        if not ans_labels:
            current_topic = explicit_topic or question
            continue

        extracted.append(
            {
                "row_no": ridx,
                "tt": tt,
                "topic": topic,
                "question": question,
                "A": _normalize(row[c_a] if c_a is not None and c_a < len(row) else ""),
                "B": _normalize(row[c_b] if c_b is not None and c_b < len(row) else ""),
                "C": _normalize(row[c_c] if c_c is not None and c_c < len(row) else ""),
                "D": _normalize(row[c_d] if c_d is not None and c_d < len(row) else ""),
                "ans": ans_labels,
                "subject_code": subject_codes[0],
                "subject_codes": subject_codes,
                "rating_explicit": rating_explicit,
            }
        )
    return extracted


def _read_xlsx_records(path: Path, default_subject_code: str):
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise ImportError(f"Khong the doc file Excel: {exc}") from exc

    try:
        ws = wb.active
        if ws is None:
            raise ImportError("File Excel khong co worksheet.")

        # Some Excel exporters omit the worksheet dimension metadata. In
        # read-only mode openpyxl then exposes max_row as None until it scans
        # the sheet explicitly.
        if ws.max_row is None:
            try:
                ws.calculate_dimension(force=True)
            except (TypeError, UnboundLocalError, ValueError) as exc:
                raise ImportError("Worksheet rong hoac khong xac dinh duoc vung du lieu.") from exc

        if not ws.max_row:
            raise ImportError("Worksheet khong co du lieu.")

        rows = list(
            ws.iter_rows(
                min_row=1,
                max_row=min(ws.max_row, 5000),
                min_col=1,
                max_col=20,
                values_only=True,
            )
        )
    finally:
        wb.close()
    return _extract_from_table_rows(rows, default_subject_code)


def _read_csv_records(path: Path, default_subject_code: str):
    rows = _read_csv(path)
    records = []
    for i, row in enumerate(rows, start=2):
        code = _normalize(
            _csv_field(row, "Mã câu hỏi", "Ma cau hoi", "TT", "STT", "code")
        )
        question = _normalize(
            _csv_field(
                row,
                "Nội dung câu hỏi",
                "Noi dung cau hoi",
                "QUESTION",
                "QUESTIONS",
                "Cau hoi",
                "Câu hỏi",
            )
        )
        if not code and not question:
            continue
        if not question:
            continue
        ans_labels = _normalize_answer_labels(
            _csv_field(row, "Đáp án đúng", "Dap an dung", "ANS", "ANSWER", "DAP AN")
        )
        if not ans_labels:
            continue
        topic_raw = _normalize(
            _csv_field(
                row,
                "Chủ đề",
                "Chu de",
                "Loại kiến thức",
                "Loại Kiến Thức",
                "Loai kien thuc",
                "Loại kiến thức (Chuyên môn/nghiệp vụ)",
            )
            or "Khac"
        )
        explicit_ratings = _parse_rating_values(_csv_field(row, "rating", "RATING", "NHOM", "NHÓM", "SUBJECT"))
        rating_explicit = bool(explicit_ratings)
        subject_codes = explicit_ratings or [default_subject_code]
        records.append(
            {
                "row_no": i,
                "tt": code or str(i),
                "topic": _classify_category(topic_raw, question),
                "question": question,
                "A": _normalize(_csv_field(row, "A", "1")),
                "B": _normalize(_csv_field(row, "B", "2")),
                "C": _normalize(_csv_field(row, "C", "3")),
                "D": _normalize(_csv_field(row, "D", "4")),
                "ans": ans_labels,
                "subject_code": subject_codes[0],
                "subject_codes": subject_codes,
                "rating_explicit": rating_explicit,
            }
        )
    return records


def import_questions_from_file(file_path: Path, subject_code: str):
    if not file_path.exists():
        raise ImportError(f"File not found: {file_path}")

    ext = file_path.suffix.lower()
    if ext == ".csv":
        records = _read_csv_records(file_path, subject_code)
    elif ext in (".xlsx", ".xlsm"):
        records = _read_xlsx_records(file_path, subject_code)
    else:
        raise ImportError("Only CSV/XLSX are supported")

    imported = 0
    imported_by_subject = {code: 0 for code in SUBJECT_GROUPS}
    source = _source_prefix(file_path)
    with transaction.atomic():
        for rec in records:
            targets = _resolve_subject_targets(
                file_path,
                rec.get("subject_codes") or [rec["subject_code"]],
                rec.get("rating_explicit", False),
            )
            for target_code in targets:
                subject, _ = Subject.objects.get_or_create(code=target_code, defaults={"name": target_code})
                category, _ = Category.objects.get_or_create(subject=subject, name=rec["topic"] or "Khac")

                # Deterministic per source file, so importing several files for one subject does not overwrite another bank.
                code = f"{subject.code}-{source}-{rec['row_no']:04d}"
                question, _ = Question.objects.update_or_create(
                    code=code,
                    defaults={
                        "content": rec["question"],
                        "subject": subject,
                        "category": category,
                        "difficulty": "",
                        "topic": rec["topic"],
                        "status": Question.STATUS_DRAFT,
                        "question_type": Question.TYPE_MULTI if len(rec["ans"]) > 1 else Question.TYPE_SINGLE,
                        "reference_document": None,
                        "explanation": "",
                    },
                )

                Answer.objects.filter(question=question).delete()
                opts = [("A", rec["A"]), ("B", rec["B"]), ("C", rec["C"]), ("D", rec["D"])]
                for i, (label, text) in enumerate(opts, start=1):
                    if not text:
                        continue
                    Answer.objects.create(
                        question=question,
                        label=label,
                        content=text,
                        is_correct=label in rec["ans"],
                        order=i,
                    )
                imported += 1
                if target_code in imported_by_subject:
                    imported_by_subject[target_code] += 1

    return {
        "imported": imported,
        "record_count": len(records),
        "by_subject": {k: v for k, v in imported_by_subject.items() if v},
    }


def preview_import_classification(file_path: Path, subject_code: str):
    if not file_path.exists():
        raise ImportError(f"File not found: {file_path}")

    ext = file_path.suffix.lower()
    if ext == ".csv":
        records = _read_csv_records(file_path, subject_code)
    elif ext in (".xlsx", ".xlsm"):
        records = _read_xlsx_records(file_path, subject_code)
    else:
        raise ImportError("Only CSV/XLSX are supported")

    by_subject_category = {}
    for rec in records:
        for target_subject in _resolve_subject_targets(
            file_path,
            rec.get("subject_codes") or [rec["subject_code"]],
            rec.get("rating_explicit", False),
        ):
            key = (target_subject, rec["topic"])
            by_subject_category[key] = by_subject_category.get(key, 0) + 1

    rows = [
        {"subject": subject, "category": category, "count": count}
        for (subject, category), count in sorted(by_subject_category.items())
    ]
    return {
        "file_name": file_path.name,
        "source_code": _source_prefix(file_path),
        "record_count": len(records),
        "import_count": sum(row["count"] for row in rows),
        "rows": rows,
    }

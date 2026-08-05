"""Replace APS/ADC banks from the new Excel files and merge them into SUP.

- APS <- BỘ ĐỀ APS MỚI.xlsx (full replace)
- ADC <- BỘ ĐỀ ADC MỚI.xlsx (all sheets, full replace)
- SUP <- (existing SUP minus content duplicates of the two new banks)
         + all questions from the two new banks
"""

from __future__ import annotations

import os
import re
import shutil
import sys
import unicodedata
from pathlib import Path

ONLINE_ROOT = Path(__file__).resolve().parent.parent
REPO_ROOT = ONLINE_ROOT.parent
sys.path.insert(0, str(ONLINE_ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.sqlite_settings")

import django

django.setup()

from django.db import transaction

from exam_bank.importers import ImportError, _read_xlsx_records, import_questions_from_file
from exam_bank.models import Answer, Category, Question, Subject
from exam_bank.question_lifecycle import archive_question, question_has_history
from exam_bank.question_selection import is_tsn_question


APS_SOURCE = REPO_ROOT / "BỘ ĐỀ APS MỚI.xlsx"
ADC_SOURCE = REPO_ROOT / "BỘ ĐỀ ADC MỚI.xlsx"
IMPORTS_DIR = ONLINE_ROOT / "imports"
APS_IMPORT = IMPORTS_DIR / "BỘ ĐỀ APS MỚI.xlsx"
ADC_IMPORT = IMPORTS_DIR / "BỘ ĐỀ ADC MỚI.xlsx"


def _plain(value: str) -> str:
    text = str(value or "").replace("đ", "d").replace("Đ", "D")
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", text).strip().upper()


def _retire_questions(queryset):
    deleted = 0
    archived = 0
    for question in queryset.iterator(chunk_size=200):
        if question_has_history(question):
            archive_question(question)
            archived += 1
        else:
            question.delete()
            deleted += 1
    return deleted, archived


def _content_keys_from_file(path: Path, default_subject: str) -> set[str]:
    records = _read_xlsx_records(path, default_subject)
    return {_plain(rec["question"]) for rec in records if rec.get("question")}


def _force_dual_rating(path: Path, primary: str) -> None:
    """Rewrite RATING cells so each row imports into primary + SUP."""
    from openpyxl import load_workbook

    wb = load_workbook(path)
    target = f"{primary},SUP"
    for ws in wb.worksheets:
        header_idx = None
        rating_col = None
        question_col = None
        max_col = 0
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, max_col=20), start=1):
            keys = []
            for j, cell in enumerate(row, start=1):
                key = _plain(cell.value)
                if not key:
                    continue
                keys.append((j, key))
                max_col = max(max_col, j)
                if key == "RATING":
                    rating_col = j
                if (
                    key == "QUESTION"
                    or key.startswith("NOI DUNG CAU HOI")
                    or ("CAU HOI" in key and "PHUONG AN" not in key)
                ):
                    question_col = j
            key_set = {key for _, key in keys}
            has_question = any(
                key == "QUESTION"
                or key.startswith("NOI DUNG CAU HOI")
                or ("CAU HOI" in key and "PHUONG AN" not in key)
                for key in key_set
            )
            has_ans = any(key == "ANS" or key.startswith("DAP AN") for key in key_set)
            if has_question and has_ans:
                header_idx = i
                if rating_col is None:
                    rating_col = max_col + 1
                    ws.cell(row=i, column=rating_col, value="RATING")
                break
        if header_idx is None or rating_col is None or question_col is None:
            continue
        for ridx in range(header_idx + 1, (ws.max_row or header_idx) + 1):
            if not ws.cell(row=ridx, column=question_col).value:
                continue
            ws.cell(row=ridx, column=rating_col, value=target)
    wb.save(path)
    wb.close()


def _prepare_import_files() -> None:
    IMPORTS_DIR.mkdir(parents=True, exist_ok=True)
    if not APS_SOURCE.exists():
        raise SystemExit(f"Missing source file: {APS_SOURCE}")
    if not ADC_SOURCE.exists():
        raise SystemExit(f"Missing source file: {ADC_SOURCE}")
    shutil.copy2(APS_SOURCE, APS_IMPORT)
    shutil.copy2(ADC_SOURCE, ADC_IMPORT)
    _force_dual_rating(APS_IMPORT, "APS")
    _force_dual_rating(ADC_IMPORT, "ADC")


def _dedupe_sup_by_content() -> int:
    """Keep one SUP question per normalized content; prefer TSN-marked copy."""
    groups: dict[str, list[Question]] = {}
    qs = Question.objects.filter(subject__code="SUP").exclude(
        status=Question.STATUS_LOCKED
    )
    for question in qs:
        groups.setdefault(_plain(question.content), []).append(question)

    retired = 0
    for group in groups.values():
        group.sort(
            key=lambda item: (
                0 if is_tsn_question(item) else 1,
                item.id,
            )
        )
        for extra in group[1:]:
            if question_has_history(extra):
                archive_question(extra)
            else:
                extra.delete()
            retired += 1
    return retired


def main() -> None:
    print("Preparing import copies with APS/ADC + SUP ratings...")
    _prepare_import_files()

    aps_keys = _content_keys_from_file(APS_IMPORT, "APS")
    adc_keys = _content_keys_from_file(ADC_IMPORT, "ADC")
    new_keys = aps_keys | adc_keys
    print(f"New unique contents: APS={len(aps_keys)} ADC={len(adc_keys)} union={len(new_keys)}")

    with transaction.atomic():
        aps_del, aps_arch = _retire_questions(
            Question.objects.filter(subject__code="APS").exclude(
                status=Question.STATUS_LOCKED
            )
        )
        adc_del, adc_arch = _retire_questions(
            Question.objects.filter(subject__code="ADC").exclude(
                status=Question.STATUS_LOCKED
            )
        )
        print(f"Retired APS: deleted={aps_del} archived={aps_arch}")
        print(f"Retired ADC: deleted={adc_del} archived={adc_arch}")

        sup_dup_qs = (
            Question.objects.filter(subject__code="SUP")
            .exclude(status=Question.STATUS_LOCKED)
            .iterator(chunk_size=200)
        )
        sup_removed = 0
        for question in sup_dup_qs:
            if _plain(question.content) in new_keys:
                if question_has_history(question):
                    archive_question(question)
                else:
                    question.delete()
                sup_removed += 1
        print(f"Removed SUP duplicates vs new banks: {sup_removed}")

        try:
            aps_result = import_questions_from_file(APS_IMPORT, "APS")
            adc_result = import_questions_from_file(ADC_IMPORT, "ADC")
        except ImportError as exc:
            raise SystemExit(f"Import failed: {exc}") from exc

        print("Import APS:", aps_result)
        print("Import ADC:", adc_result)

        # Approve newly imported (and remaining active) APS/ADC/SUP questions.
        for code in ("APS", "ADC", "SUP"):
            updated = (
                Question.objects.filter(subject__code=code)
                .exclude(status=Question.STATUS_LOCKED)
                .update(status=Question.STATUS_APPROVED)
            )
            print(f"Approved {code}: {updated}")

        sup_deduped = _dedupe_sup_by_content()
        print(f"SUP intra-bank content dedupe retired: {sup_deduped}")

        # Prefer/repair TSN markers on SUP, and share ADC TSN sheet into APS so
        # 100-question papers can still hit the 35% TSN quota.
        _upgrade_sup_tsn_topics()
        copied = _copy_adc_tsn_into_subjects(("APS", "SUP"))
        print(f"Copied ADC TSN sheet rows into APS/SUP: {copied}")

    for code in ("APS", "ADC", "SUP"):
        qs = list(
            Question.objects.filter(
                subject__code=code, status=Question.STATUS_APPROVED
            ).exclude(is_locked_for_official_exam=True)
        )
        tsn = sum(1 for q in qs if is_tsn_question(q))
        print(
            f"{code}: approved={len(qs)} tsn={tsn} "
            f"pct={round(100 * tsn / len(qs), 1) if qs else 0}"
        )
        answers = Answer.objects.filter(question__in=qs).count()
        print(f"{code}: answers={answers}")


def _upgrade_sup_tsn_topics() -> None:
    adc_tsn_by_key = {
        _plain(question.content): question
        for question in Question.objects.filter(
            subject__code="ADC",
            status=Question.STATUS_APPROVED,
            topic__icontains="Tan Son",
        )
    }
    for question in Question.objects.filter(subject__code="SUP").exclude(
        status=Question.STATUS_LOCKED
    ):
        src = adc_tsn_by_key.get(_plain(question.content))
        if src is None or is_tsn_question(question):
            continue
        question.topic = src.topic
        question.save(update_fields=["topic", "updated_at"])


def _copy_adc_tsn_into_subjects(subject_codes) -> dict:
    adc_tsn = list(
        Question.objects.filter(
            subject__code="ADC",
            status=Question.STATUS_APPROVED,
            topic__icontains="Tan Son",
        ).prefetch_related("answers")
    )
    created = {code: 0 for code in subject_codes}
    for code in subject_codes:
        subject = Subject.objects.get(code=code)
        keys = {
            _plain(question.content)
            for question in Question.objects.filter(
                subject__code=code,
                status=Question.STATUS_APPROVED,
                is_locked_for_official_exam=False,
            )
        }
        for src in adc_tsn:
            key = _plain(src.content)
            if key in keys:
                continue
            category, _ = Category.objects.get_or_create(
                subject=subject,
                name=src.category.name if src.category_id else "Khac",
            )
            question = Question.objects.create(
                code=f"{code}-FROM-ADC-TSN-{src.id}",
                content=src.content,
                subject=subject,
                category=category,
                topic=src.topic,
                status=Question.STATUS_APPROVED,
                question_type=src.question_type,
            )
            for order, answer in enumerate(
                src.answers.all().order_by("order", "id"), start=1
            ):
                Answer.objects.create(
                    question=question,
                    label=answer.label,
                    content=answer.content,
                    is_correct=answer.is_correct,
                    order=order,
                )
            keys.add(key)
            created[code] += 1
    return created


if __name__ == "__main__":
    main()
